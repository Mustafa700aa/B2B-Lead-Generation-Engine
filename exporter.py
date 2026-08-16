"""
Data Exporter module for the B2B Lead Generation Engine.
Uses Pandas and OpenPyXL to clean, standardize, and export leads into
professionally styled Excel (.xlsx) reports.
"""

import re
from datetime import datetime
from pathlib import Path
from typing import List, Optional
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import (
    EXPORTS_DIR,
    EXCEL_SHEET_NAME,
    EXCEL_HEADER_BG_COLOR,
    EXCEL_HEADER_FONT_COLOR,
    EXCEL_ROW_ALT_BG_COLOR,
)
from models import Lead
from database import DatabaseManager
from logger import get_logger

logger = get_logger("DataExporter")


class DataExporter:
    """
    Pandas-powered data transformation and styled Excel export engine.
    """

    def __init__(self, db: Optional[DatabaseManager] = None):
        self.db = db or DatabaseManager()

    def _normalize_phone_number(self, phone: Optional[str]) -> str:
        """
        Clean and standardize phone number formatting.
        Preserves international country codes (e.g. +971).
        """
        if not phone or pd.isna(phone):
            return "N/A"

        phone_str = str(phone).strip()
        
        # Remove unwanted label text if present
        phone_str = re.sub(r"(?i)phone:\s*", "", phone_str)
        # Collapse multiple spaces
        phone_str = re.sub(r"\s+", " ", phone_str).strip()

        return phone_str if phone_str else "N/A"

    def _clean_leads_dataframe(self, leads: List[Lead]) -> pd.DataFrame:
        """
        Convert Lead objects to a cleaned, formatted Pandas DataFrame.
        """
        if not leads:
            return pd.DataFrame()

        data = [lead.to_dict() for lead in leads]
        df = pd.DataFrame(data)

        # 1. Deduplication
        df.drop_duplicates(subset=["name", "website"], keep="first", inplace=True)
        if "phone" in df.columns:
            df.drop_duplicates(subset=["name", "phone"], keep="first", inplace=True)

        # 2. Column Renaming for professional presentation
        column_mapping = {
            "id": "Lead ID",
            "name": "Company Name",
            "phone": "Phone Number",
            "email": "Contact Email",
            "website": "Website URL",
            "address": "Full Address",
            "rating": "Google Rating",
            "reviews_count": "Total Reviews",
            "query": "Search Query",
            "maps_url": "Google Maps URL",
            "is_enriched": "Enriched",
            "created_at": "Discovered At",
        }

        # Select and reorder desired columns
        desired_columns = [
            "id", "name", "phone", "email", "website", "address",
            "rating", "reviews_count", "query", "maps_url", "is_enriched", "created_at"
        ]
        present_columns = [col for col in desired_columns if col in df.columns]
        df = df[present_columns].rename(columns=column_mapping)

        # 3. Clean and standardize fields
        if "Phone Number" in df.columns:
            df["Phone Number"] = df["Phone Number"].apply(self._normalize_phone_number)

        if "Contact Email" in df.columns:
            df["Contact Email"] = df["Contact Email"].fillna("N/A").replace("", "N/A")

        if "Website URL" in df.columns:
            df["Website URL"] = df["Website URL"].fillna("N/A").replace("", "N/A")

        if "Full Address" in df.columns:
            df["Full Address"] = df["Full Address"].fillna("N/A").replace("", "N/A")

        if "Google Rating" in df.columns:
            df["Google Rating"] = pd.to_numeric(df["Google Rating"], errors="coerce").fillna(0.0)

        if "Total Reviews" in df.columns:
            df["Total Reviews"] = pd.to_numeric(df["Total Reviews"], errors="coerce").fillna(0).astype(int)

        if "Enriched" in df.columns:
            df["Enriched"] = df["Enriched"].apply(lambda x: "Yes" if bool(x) else "No")

        return df

    def _style_excel_workbook(self, file_path: Path) -> None:
        """
        Apply professional typography, dark slate headers, zebra striping,
        borders, and auto-fit column widths to the Excel worksheet.
        """
        wb = load_workbook(file_path)
        ws = wb.active
        ws.title = EXCEL_SHEET_NAME

        # Style Definitions
        header_font = Font(name="Segoe UI", size=11, bold=True, color=EXCEL_HEADER_FONT_COLOR)
        header_fill = PatternFill(start_color=EXCEL_HEADER_BG_COLOR, end_color=EXCEL_HEADER_BG_COLOR, fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        data_font = Font(name="Segoe UI", size=10)
        email_font = Font(name="Segoe UI", size=10, bold=True, color="002060")
        link_font = Font(name="Segoe UI", size=10, color="0563C1", underline="single")
        
        alt_fill = PatternFill(start_color=EXCEL_ROW_ALT_BG_COLOR, end_color=EXCEL_ROW_ALT_BG_COLOR, fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )

        # 1. Format Header Row
        ws.row_dimensions[1].height = 28
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 2. Format Data Rows
        email_col_idx = None
        web_col_idx = None
        maps_col_idx = None

        for col_idx in range(1, ws.max_column + 1):
            col_name = str(ws.cell(row=1, column=col_idx).value or "")
            if "email" in col_name.lower():
                email_col_idx = col_idx
            elif "website" in col_name.lower():
                web_col_idx = col_idx
            elif "maps" in col_name.lower():
                maps_col_idx = col_idx

        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 20
            is_alt = (row_idx % 2 == 0)
            row_fill = alt_fill if is_alt else white_fill

            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                cell.fill = row_fill
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

                # Highlight Emails
                if col_idx == email_col_idx and cell.value and cell.value != "N/A":
                    cell.font = email_font

                # Hyperlink Websites & Maps URLs
                if col_idx in (web_col_idx, maps_col_idx) and cell.value and str(cell.value).startswith("http"):
                    cell.hyperlink = cell.value
                    cell.font = link_font

        # 3. Auto-fit Column Widths
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or "")
                if len(val) > max_len:
                    max_len = len(val)
            adjusted_width = min(max(max_len + 4, 12), 45)
            ws.column_dimensions[col_letter].width = adjusted_width

        # 4. Freeze header row & Enable AutoFilter
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        wb.save(file_path)
        logger.info(f"Excel workbook styled successfully: {file_path}")

    def export_to_excel(
        self,
        leads: Optional[List[Lead]] = None,
        output_filename: Optional[str] = None
    ) -> Path:
        """
        Export leads to an aesthetically formatted Excel report.

        :param leads: Optional list of Lead objects. If None, queries all leads from DB.
        :param output_filename: Optional custom filename.
        :return: Path to generated Excel file
        """
        target_leads = leads if leads is not None else self.db.get_all_leads()

        if not target_leads:
            logger.warning("No leads found in database to export.")
            raise ValueError("No leads available to export.")

        df = self._clean_leads_dataframe(target_leads)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = output_filename or f"b2b_leads_{timestamp}.xlsx"
        if not filename.endswith(".xlsx"):
            filename += ".xlsx"

        file_path = EXPORTS_DIR / filename

        # Write initial dataframe to Excel
        df.to_excel(file_path, index=False, engine="openpyxl")
        
        # Apply OpenPyXL formatting & styling
        self._style_excel_workbook(file_path)

        logger.info(f"Export completed. {len(df)} leads written to: {file_path}")
        return file_path
